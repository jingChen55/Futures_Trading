





#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PTA期权链API模块
提供期权链T型报价、希腊字母、隐波、PCR等数据

数据源：
- akshare: 日频历史数据
- tqsdk: 实时数据（需要认证）
"""

from __future__ import annotations
import json
import math
import re
import sqlite3
import threading
import time
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass, asdict
from scipy.stats import norm
import pandas as pd
import numpy as np
import akshare as ak
import os

# 天勤量化 TqSdk
from tqsdk import TqApi, TqAuth
TQS_USER = os.environ.get('TQS_AUTH_USER', 'mingmingliu')
TQS_PASS = os.environ.get('TQS_AUTH_PASS', 'Liuzhaoning2025')

# ==================== 常量定义 ====================

# CZCE PTA期权合约代码到实际到期日的映射
# 注意: TA606只是代号，不代表实际到期月份
# 根据akshare数据(20260416)，TA606是近月合约，实际到期日为5月13日
# CZCE规则: 期权到期日 = 标的期货合约交割月前一个月的第5个交易日
PTA_EXPIRY_MAP = {
    'TA606': '20260513',  # 2026年5月13日到期 (近月)
    'TA607': '20260615',  # 2026年6月15日到期
    'TA608': '20260714',  # 2026年7月14日到期
    'TA609': '20260812',  # 2026年8月12日到期
    'TA610': '20260914',  # 2026年9月14日到期
    'TA611': '20261014',  # 2026年10月14日到期
    'TA701': '20270113',  # 2027年1月13日到期
}

# 期权合约到标的期货的映射（akshare的symbol命名）
# TA606期权 -> TA606期货, TA609期权 -> TA609期货
OPTION_TO_FUTURES_SYMBOL = {
    'TA606': 'TA606',
    'TA607': 'TA607',
    'TA608': 'TA608',
    'TA609': 'TA609',
    'TA610': 'TA610',
    'TA611': 'TA611',
    'TA701': 'TA701',
}

def get_expiry_date(expiry_code: str) -> str:
    """获取合约的实际到期日"""
    return PTA_EXPIRY_MAP.get(expiry_code, '20260615')  # 默认6月

def get_futures_symbol(expiry_code: str) -> str:
    """获取期权合约对应的标的期货akshare symbol"""
    return OPTION_TO_FUTURES_SYMBOL.get(expiry_code, 'TA606')

def sort_expiry_list(expiry_list: List) -> List:
    """按实际到期日排序（最近优先）"""
    return sorted(expiry_list, key=lambda x: get_expiry_date(x.expiry))

# PCR阈值
PCR_LOUD_THRESHOLD = 0.8   # 偏多阈值
PCR_BEAR_THRESHOLD = 1.0   # 偏空阈值

# 持仓变化阈值 (%)
OI_CHANGE_NOISE = 0.10      # 噪音区间
OI_CHANGE_SIGNIFICANT = 0.25 # 显著变化

# 隐波变化阈值 (按波环境)
IV_CHANGE_LOW = 1.5       # 低波环境(<20%)
IV_CHANGE_MID = 2.0       # 中波环境(20-30%)
IV_CHANGE_HIGH = 3.0      # 高波环境(>30%)

# 希腊字母计算用
RISK_FREE_RATE = 0.03

# ==================== 数据结构 ====================

@dataclass
class ExpiryData:
    """到期日数据"""
    expiry: str           # 合约代码 TA606等
    actual_expiry_date: str  # 实际到期日 20260615
    trading_date: str      # 交易日期
    call_count: int        # 认购合约数
    put_count: int         # 认沽合约数
    total_volume_call: int
    total_volume_put: int
    total_oi_call: int
    total_oi_put: int
    volume_pcr: float    # 成交PCR
    position_pcr: float   # 持仓PCR

@dataclass
class StrikeRow:
    """T型报价单行"""
    expiry: str       # 合约代码 TA606等
    actual_expiry_date: str  # 实际到期日 20260615
    strike: float
    # Call边
    call_code: str
    call_price: float      # 今结算(结算价)
    call_close: float      # 今收盘(成交价)
    call_settlement: float # 昨结算
    call_change1: float    # 涨跌1(今收盘-昨结算)
    call_change2: float    # 涨跌2(今结算-昨结算)
    call_iv: float
    call_iv_change: float
    call_delta: float
    call_gamma: float
    call_theta: float
    call_vega: float
    call_volume: int
    call_volume_change: int  # 成交量变化（绝对值）
    call_oi: int
    call_oi_change: float
    call_exercise: int      # 行权量
    # Put边
    put_code: str
    put_price: float       # 今结算(结算价)
    put_close: float        # 今收盘(成交价)
    put_settlement: float  # 昨结算
    put_change1: float      # 涨跌1(今收盘-昨结算)
    put_change2: float      # 涨跌2(今结算-昨结算)
    put_iv: float
    put_iv_change: float
    put_delta: float
    put_gamma: float
    put_theta: float
    put_vega: float
    put_volume: int
    put_volume_change: int  # 成交量变化（绝对值）
    put_oi: int
    put_oi_change: float
    put_exercise: int
    # 前日IV（用于IV曲线日间比较）
    prev_call_iv: float = 0.0
    prev_put_iv: float = 0.0       # 行权量

# ==================== 希腊字母计算 ====================

def calculate_days_to_expiry(expiry: str, trade_date: str = None) -> float:
    """计算剩余到期时间(年)
    
    Args:
        expiry: 到期日 YYYYMMDD 或期权代码如 'TA606'
        trade_date: 交易日期 YYYYMMDD，默认今日
    """
    if trade_date is None:
        trade_date = datetime.now().strftime('%Y%m%d')
    
    # 如果expiry是期权代码格式（如TA606），转换为实际日期
    if expiry and len(expiry) == 5 and expiry[:2].isalpha():
        expiry = get_expiry_date(expiry)
    
    try:
        expiry_dt = datetime.strptime(expiry, '%Y%m%d')
        trade_dt = datetime.strptime(trade_date, '%Y%m%d')
        days = (expiry_dt - trade_dt).days
        return max(days / 365.0, 1/365)  # 至少1天
    except:
        return 0.25  # 默认3个月

def bs_price(S: float, K: float, T: float, r: float, sigma: float, option_type: str) -> float:
    """Black-Scholes期权定价"""
    if T <= 0 or sigma <= 0:
        return 0.0
    
    d1 = (math.log(S / K) + (r + 0.5 * sigma ** 2) * T) / (sigma * math.sqrt(T))
    d2 = d1 - sigma * math.sqrt(T)
    
    if option_type == 'C':
        price = S * norm.cdf(d1) - K * math.exp(-r * T) * norm.cdf(d2)
    else:
        price = K * math.exp(-r * T) * norm.cdf(-d2) - S * norm.cdf(-d1)
    
    return price

def calculate_iv(option_price: float, S: float, K: float, T: float, r: float, option_type: str, 
                  tol: float = 1e-6, max_iter: int = 100) -> float:
    """从期权价格反算隐含波动率 (二分法)
    
    Args:
        option_price: 期权市场价格
        S: 标的价格
        K: 行权价
        T: 到期时间(年)
        r: 无风险利率
        option_type: 'C' 或 'P'
        tol: 收敛容差
        max_iter: 最大迭代次数
        
    Returns:
        float: 隐含波动率 (如计算失败返回0)
    """
    if T <= 0 or option_price <= 0:
        return 0.0
    
    # 检查边界条件
    intrinsic = max(S - K, 0) if option_type == 'C' else max(K - S, 0)
    if option_price < intrinsic:
        return 0.0  # 价格低于内在价值，无效
    
    # 二分法查找IV
    sigma_low = 0.001
    sigma_high = 5.0  # 500%的波动率上限
    
    price_low = bs_price(S, K, T, r, sigma_low, option_type)
    price_high = bs_price(S, K, T, r, sigma_high, option_type)
    
    # 如果价格不在合理范围内，返回0
    if option_price < price_low or option_price > price_high:
        return 0.0
    
    for _ in range(max_iter):
        sigma_mid = (sigma_low + sigma_high) / 2
        price_mid = bs_price(S, K, T, r, sigma_mid, option_type)
        
        if abs(price_mid - option_price) < tol:
            break
        
        if price_mid < option_price:
            sigma_low = sigma_mid
        else:
            sigma_high = sigma_mid
    
    return sigma_mid

def estimate_underlying_from_options(df: pd.DataFrame, expiry: str, r: float = 0.03) -> float:
    """用Put-Call Parity从期权数据估算标的期货价格
    
    Args:
        df: 期权数据DataFrame
        expiry: 到期月份代码，如 'TA606'
        r: 无风险利率
        
    Returns:
        float: 估算的标的期货价格（如估算失败返回0）
    """
    try:
        # 筛选指定月份的期权
        expiry_df = df[df['合约代码'].str.startswith(expiry)]
        if len(expiry_df) == 0:
            return 0.0
        
        # 计算每个行权价的成交量
        expiry_df = expiry_df.copy()
        expiry_df['strike'] = expiry_df['合约代码'].apply(lambda x: int(x[-4:]) if x[-4:].isdigit() else 0)
        expiry_df['volume'] = expiry_df.get('成交量(手)', 0).fillna(0)
        
        # 找成交量最大的行权价（作为ATM参考）
        if len(expiry_df) == 0:
            return 0.0
        
        max_vol_strike = expiry_df.loc[expiry_df['volume'].idxmax(), 'strike']
        
        # 找ATM附近的call和put
        atm_calls = expiry_df[(expiry_df['合约代码'].str.contains('C')) & 
                              (abs(expiry_df['strike'] - max_vol_strike) <= 100)]
        atm_puts = expiry_df[(expiry_df['合约代码'].str.contains('P')) & 
                             (abs(expiry_df['strike'] - max_vol_strike) <= 100)]
        
        if len(atm_calls) == 0 or len(atm_puts) == 0:
            return 0.0
        
        # 取ATM call和put（用成交量加权）
        call = atm_calls.loc[atm_calls['成交量(手)'].idxmax()]
        put = atm_puts.loc[atm_puts['成交量(手)'].idxmax()]
        
        K = float(call['strike'])
        C = float(call.get('今结算', 0) or call.get('今收盘', 0))
        P = float(put.get('今结算', 0) or put.get('今收盘', 0))
        
        if C <= 0 or P <= 0:
            return 0.0
        
        # 计算到期时间
        T = calculate_days_to_expiry(expiry) / 365.0
        if T <= 0:
            T = 0.1  # 默认10天
        
        # Put-Call Parity: C - P = S - K*exp(-rT)
        # => S = C - P + K*exp(-rT)
        S = C - P + K * math.exp(-r * T)
        
        # 合理性检查：S应该在[K-20%, K+20%]范围内
        if S < K * 0.8 or S > K * 1.2:
            return 0.0
        
        return S
    except:
        return 0.0

def get_tq_futures_price(symbol: str = 'KQ.m@CZCE.TA', timeout: float = 5.0) -> float:
    """从TqSdk获取期货实时价格（主力合约）
    
    Args:
        symbol: 期货合约代码，默认 'KQ.m@CZCE.TA' 主力合约
        timeout: 超时秒数
        
    Returns:
        float: 最新价，失败返回0
    """
    try:
        import threading
        result = {'price': 0, 'done': False}
        
        def fetch():
            try:
                api = TqApi(auth=TqAuth(TQS_USER, TQS_PASS), debug=False)
                quote = api.get_quote(symbol)
                start = time.time()
                while time.time() - start < timeout and not result['done']:
                    try:
                        api.wait_update(deadline=min(time.time() + 1.0, start + timeout))
                        if quote.last_price and quote.last_price > 0:
                            result['price'] = float(quote.last_price)
                            break
                    except Exception:
                        break
                try:
                    api.close()
                except Exception:
                    pass
            except Exception:
                pass
            finally:
                result['done'] = True
        
        t = threading.Thread(target=fetch, daemon=True)
        t.start()
        t.join(timeout=timeout + 2)
        return result['price']
    except Exception:
        return 0


def get_tq_ta606_price(timeout: float = 5.0) -> float:
    """从TqSdk获取TA606（PTA 6月期货）实时价格
    
    Args:
        timeout: 超时秒数
        
    Returns:
        float: TA606最新价，失败返回0
    """
    try:
        import threading
        result = {'price': 0, 'done': False}
        
        def fetch():
            try:
                api = TqApi(auth=TqAuth(TQS_USER, TQS_PASS), debug=False)
                # TA606合约代码: CZCE.TA606
                quote = api.get_quote('CZCE.TA606')
                start = time.time()
                while time.time() - start < timeout and not result['done']:
                    try:
                        api.wait_update(deadline=min(time.time() + 1.0, start + timeout))
                        if quote.last_price and quote.last_price > 0:
                            result['price'] = float(quote.last_price)
                            break
                    except Exception:
                        break
                try:
                    api.close()
                except Exception:
                    pass
            except Exception:
                pass
            finally:
                result['done'] = True
        
        t = threading.Thread(target=fetch, daemon=True)
        t.start()
        t.join(timeout=timeout + 2)
        return result['price']
    except Exception:
        return 0

def get_tq_option_prices(symbols: List[str]) -> Dict[str, Dict]:
    """从TqSdk获取期权实时价格（暂时禁用，因TqSdk连接阻塞）
    
    Args:
        symbols: 期权合约代码列表，如 ['CZCE.TA606C6800', 'CZCE.TA606P6800']
        
    Returns:
        dict: {symbol: {last_price, volume, open_interest, bid_price, ask_price, ...}}
    """
    # TqSdk暂时禁用，因连接会阻塞
    return {}

def calculate_greeks(S: float, K: float, T: float, r: float, sigma: float, option_type: str) -> Dict[str, float]:
    """计算希腊字母
    
    Args:
        S: 标的价格
        K: 行权价
        T: 到期时间(年)
        r: 无风险利率
        sigma: 波动率
        option_type: 'C' 或 'P'
        
    Returns:
        dict: {delta, gamma, theta, vega}
    """
    if T <= 0 or sigma <= 0:
        return {'delta': 0, 'gamma': 0, 'theta': 0, 'vega': 0}
    
    sqrtT = math.sqrt(T)
    d1 = (math.log(S / K) + (r + 0.5 * sigma ** 2) * T) / (sigma * sqrtT)
    d2 = d1 - sigma * sqrtT
    
    if option_type == 'C':
        delta = norm.cdf(d1)
        theta = (-S * norm.pdf(d1) * sigma / (2 * sqrtT) 
                 - r * K * math.exp(-r * T) * norm.cdf(d2)) / 365
    else:
        delta = norm.cdf(d1) - 1
        theta = (-S * norm.pdf(d1) * sigma / (2 * sqrtT) 
                 + r * K * math.exp(-r * T) * norm.cdf(-d2)) / 365
    
    gamma = norm.pdf(d1) / (S * sigma * sqrtT)
    vega = S * sqrtT * norm.pdf(d1) / 100  # 每1%波动率
    
    return {
        'delta': delta,
        'gamma': gamma,
        'theta': theta,
        'vega': vega
    }

def parse_option_code(code: str) -> Tuple[Optional[str], Optional[str], Optional[float]]:
    """解析期权合约代码
    
    Args:
        code: 合约代码，如 'TA605C4100'
        
    Returns:
        (expiry, option_type, strike)
    """
    # 匹配模式: TA605C4100 -> expiry=TA605, type=C, strike=4100
    m = re.match(r'([A-Z]+)(\d+)([CP])(\d+)', code)
    if m:
        symbol = m.group(1)
        month = m.group(2)
        option_type = m.group(3)
        strike = float(m.group(4))
        expiry = symbol + month
        return expiry, option_type, strike
    return None, None, None

def get_iv_from_data(row: pd.Series) -> float:
    """从akshare数据行获取IV"""
    iv = row.get('隐含波动率')
    if iv is None or pd.isna(iv):
        return 0.0
    return float(iv) / 100  # akshare返回的是百分比

def get_delta_from_data(row: pd.Series) -> float:
    """从akshare数据行获取Delta"""
    delta = row.get('DELTA')
    if delta is None or pd.isna(delta):
        return 0.0
    return float(delta)

# ==================== 数据获取 ====================

class AkshareOptionData:
    """使用akshare获取期权数据"""
    
    @staticmethod
    def get_recent_trade_date(days: int = 15) -> str:
        """获取最近交易日（跳过周末）"""
        today = datetime.now()
        weekday = today.weekday()  # 0=Monday, 6=Sunday
        
        # 如果今天是周末，从上周五开始
        if weekday == 6:  # Sunday
            start_offset = 2  # 从上周五开始
        elif weekday == 5:  # Saturday
            start_offset = 1  # 从昨天(周五)开始
        else:
            start_offset = 1  # 从昨天开始
        
        for i in range(start_offset, days):
            dt = today - timedelta(days=i)
            date_str = dt.strftime('%Y%m%d')
            try:
                df = ak.option_hist_czce(symbol='PTA期权', trade_date=date_str)
                if df is not None and len(df) > 0:
                    return date_str
            except:
                continue
        return None
    
    @staticmethod
    def get_option_data(trade_date: str = None) -> pd.DataFrame:
        """获取PTA期权数据
        
        Args:
            trade_date: 交易日期 YYYYMMDD，默认最近交易日
        """
        if trade_date is None:
            trade_date = AkshareOptionData.get_recent_trade_date()
            if trade_date is None:
                return pd.DataFrame()
        
        try:
            df = ak.option_hist_czce(symbol='PTA期权', trade_date=trade_date)
            if df is None or len(df) == 0:
                return pd.DataFrame()
            
            # 解析合约代码
            parsed = df['合约代码'].apply(lambda x: parse_option_code(str(x)))
            df['expiry'] = parsed.apply(lambda x: x[0] if x else None)
            df['option_type'] = parsed.apply(lambda x: x[1] if x else None)
            df['strike'] = parsed.apply(lambda x: x[2] if x else None)
            
            # 过滤无效数据
            df = df[df['option_type'].isin(['C', 'P'])]
            df = df[df['strike'].notna()]
            
            return df
        except Exception as e:
            print(f"akshare get_option_data error: {e}")
            return pd.DataFrame()

# ==================== 期权链分析 ====================

class OptionChainAnalyzer:
    """期权链分析器"""
    
    def __init__(self, underlying_price: float = 0):
        self.underlying_price = underlying_price
        self.risk_free_rate = RISK_FREE_RATE
        self.trade_date = datetime.now().strftime('%Y%m%d')
    
    def build_t_type_quote(self, df: pd.DataFrame, 
                           prev_df: pd.DataFrame = None) -> Tuple[List[ExpiryData], List[StrikeRow]]:
        """构建T型报价 - 按到期日分组，每组独立构建
        
        Args:
            df: 当前日期的期权数据
            prev_df: 昨日数据
            
        Returns:
            (expiry_list, strike_rows)
        """
        if df is None or len(df) == 0:
            return [], []
        
        # 按到期日分组计算PCR
        expiry_groups = df.groupby('expiry')
        expiry_list = []
        all_strike_rows = []
        
        # 获取天勤实时期权价格并计算IV
        tq_prices = {}
        try:
            # 构建天勤合约代码列表
            tq_symbols = []
            for _, row in df.iterrows():
                code = row['合约代码']
                expiry = row.get('expiry', '')
                option_type = row.get('option_type', '')
                if expiry and option_type:
                    tq_sym = f'CZCE.{expiry}{option_type.upper()}{int(row.get("strike", 0))}'
                    tq_symbols.append(tq_sym)
            
            # 批量获取天勤实时价格
            if tq_symbols:
                tq_prices = get_tq_option_prices(tq_symbols)
                print(f"TqSdk获取了 {len(tq_prices)} 个期权实时价格")
        except Exception as e:
            print(f"天勤数据获取失败: {e}")
        
        # 获取昨日数据用于计算变化
        prev_dict = {}
        if prev_df is not None and len(prev_df) > 0:
            for _, row in prev_df.iterrows():
                code = row['合约代码']
                # 兼容两种列名格式
                if '今结算' in row.index or '昨结算' in row.index:
                    # akshare格式
                    close = row.get('今结算', 0) or row.get('今收盘', 0)
                    iv = get_iv_from_data(row)
                    oi = row.get('持仓量', 0)
                    volume = row.get('成交量(手)', 0)
                else:
                    # session_db格式
                    close = row.get('last_price', 0)
                    iv = row.get('iv', 0) or 0
                    oi = row.get('oi', 0) or 0
                    volume = row.get('volume', 0)
                prev_dict[code] = {
                    'close': close,
                    'iv': float(iv) if iv else 0.0,
                    'oi': int(oi) if oi else 0,
                    'volume': int(volume) if volume else 0
                }
        
        for expiry, group in expiry_groups:
            calls = group[group['option_type'] == 'C']
            puts = group[group['option_type'] == 'P']
            
            vol_call = int(calls['成交量(手)'].sum()) if len(calls) > 0 else 0
            vol_put = int(puts['成交量(手)'].sum()) if len(puts) > 0 else 0
            oi_call = int(calls['持仓量'].sum()) if len(calls) > 0 else 0
            oi_put = int(puts['持仓量'].sum()) if len(puts) > 0 else 0
            
            expiry_list.append(ExpiryData(
                expiry=expiry,
                actual_expiry_date=get_expiry_date(expiry),
                trading_date=self.trade_date,
                call_count=len(calls),
                put_count=len(puts),
                total_volume_call=vol_call,
                total_volume_put=vol_put,
                total_oi_call=oi_call,
                total_oi_put=oi_put,
                volume_pcr=round(vol_put / vol_call, 4) if vol_call > 0 else 0,
                position_pcr=round(oi_put / oi_call, 4) if oi_call > 0 else 0
            ))
            
            # 该到期日的所有行权价
            expiry_strikes = sorted(group['strike'].unique())
            
            for strike in expiry_strikes:
                call_row = group[(group['option_type'] == 'C') & (group['strike'] == strike)]
                put_row = group[(group['option_type'] == 'P') & (group['strike'] == strike)]
                
                # Call数据
                if len(call_row) > 0:
                    cr = call_row.iloc[0]
                    call_code = cr['合约代码']
                    call_price = cr.get('今结算') or 0        # 结算价
                    call_close = cr.get('今收盘') or 0        # 成交价
                    call_settlement = cr.get('昨结算') or 0    # 昨结算
                    call_change1 = cr.get('涨跌1') or 0        # 涨跌1(今收-昨结)
                    call_change2 = cr.get('涨跌2') or 0        # 涨跌2(今结-昨结)
                    call_iv = get_iv_from_data(cr)
                    call_delta = get_delta_from_data(cr)
                    call_volume = int(cr.get('成交量(手)', 0) or 0)
                    call_oi = int(cr.get('持仓量', 0) or 0)
                    call_exercise = int(cr.get('行权量', 0) or 0)
                    
                    # 尝试用天勤实时价格计算IV
                    tq_sym = f'CZCE.{expiry}C{int(strike)}'
                    tq_data = tq_prices.get(tq_sym, {})
                    tq_price = tq_data.get('last_price')
                    
                    if tq_price and self.underlying_price > 0:
                        # 用天勤价格反算IV
                        T = calculate_days_to_expiry(expiry, self.trade_date)
                        tq_iv = calculate_iv(tq_price, self.underlying_price, strike, T, self.risk_free_rate, 'C')
                        print(f"DEBUG: tq_price={tq_price}, tq_iv={tq_iv}, call_iv set to {tq_iv * 100}")
                        if tq_iv > 0:
                            call_iv = tq_iv * 100  # 转换为百分比（0.33 -> 33%）
                            call_close = tq_price  # 用天勤实时价格替换
                    
                    # 计算希腊字母（如果数据没有）
                    if call_delta == 0 and call_iv > 0:
                        T = calculate_days_to_expiry(expiry, self.trade_date)
                        greeks = calculate_greeks(self.underlying_price, strike, T, self.risk_free_rate, call_iv, 'C')
                        call_delta = greeks['delta']
                    
                    # 计算IV变化（绝对百分比）和成交量变化（相对值）
                    prev_call = prev_dict.get(call_code, {})
                    call_iv_change = (call_iv - prev_call.get('iv', 0)) * 100 if prev_call else 0
                    # OI变化用相对值（百分比）
                    call_oi_change = ((call_oi - prev_call.get('oi', 0)) / prev_call.get('oi', 1) * 100) if prev_call and prev_call.get('oi', 0) > 0 else 0
                    # 成交量变化用相对值（百分比）
                    call_volume_change = ((call_volume - prev_call.get('volume', 0)) / prev_call.get('volume', 1) * 100) if prev_call and prev_call.get('volume', 0) > 0 else 0
                    
                    # 计算其他希腊字母
                    T = calculate_days_to_expiry(expiry, self.trade_date)
                    greeks = calculate_greeks(self.underlying_price, strike, T, self.risk_free_rate, call_iv if call_iv > 0 else 0.2, 'C')
                    call_gamma = greeks['gamma']
                    call_theta = greeks['theta']
                    call_vega = greeks['vega']
                else:
                    call_code = ''
                    call_price = call_close = call_settlement = call_change1 = call_change2 = 0
                    call_iv = call_delta = call_gamma = call_theta = call_vega = 0
                    call_volume = call_oi = call_iv_change = call_oi_change = call_volume_change = call_exercise = 0
                    prev_call = {}
                
                # Put数据
                if len(put_row) > 0:
                    pr = put_row.iloc[0]
                    put_code = pr['合约代码']
                    put_price = pr.get('今结算') or 0         # 结算价
                    put_close = pr.get('今收盘') or 0         # 成交价
                    put_settlement = pr.get('昨结算') or 0     # 昨结算
                    put_change1 = pr.get('涨跌1') or 0         # 涨跌1(今收-昨结)
                    put_change2 = pr.get('涨跌2') or 0         # 涨跌2(今结-昨结)
                    put_iv = get_iv_from_data(pr)
                    put_delta = get_delta_from_data(pr)
                    put_volume = int(pr.get('成交量(手)', 0) or 0)
                    put_oi = int(pr.get('持仓量', 0) or 0)
                    put_exercise = int(pr.get('行权量', 0) or 0)
                    
                    # 尝试用天勤实时价格计算IV
                    tq_sym = f'CZCE.{expiry}P{int(strike)}'
                    tq_data = tq_prices.get(tq_sym, {})
                    tq_price = tq_data.get('last_price')
                    
                    if tq_price and self.underlying_price > 0:
                        # 用天勤价格反算IV
                        T = calculate_days_to_expiry(expiry, self.trade_date)
                        tq_iv = calculate_iv(tq_price, self.underlying_price, strike, T, self.risk_free_rate, 'P')
                        if tq_iv > 0:
                            put_iv = tq_iv * 100  # 转换为百分比
                            put_close = tq_price  # 用天勤实时价格替换
                    
                    # 计算希腊字母
                    if put_delta == 0 and put_iv > 0:
                        T = calculate_days_to_expiry(expiry, self.trade_date)
                        greeks = calculate_greeks(self.underlying_price, strike, T, self.risk_free_rate, put_iv, 'P')
                        put_delta = greeks['delta']
                    
                    # 计算IV变化（绝对百分比）和成交量变化（相对值）
                    prev_put = prev_dict.get(put_code, {})
                    put_iv_change = (put_iv - prev_put.get('iv', 0)) * 100 if prev_put else 0
                    # OI变化用相对值（百分比）
                    put_oi_change = ((put_oi - prev_put.get('oi', 0)) / prev_put.get('oi', 1) * 100) if prev_put and prev_put.get('oi', 0) > 0 else 0
                    # 成交量变化用相对值（百分比）
                    put_volume_change = ((put_volume - prev_put.get('volume', 0)) / prev_put.get('volume', 1) * 100) if prev_put and prev_put.get('volume', 0) > 0 else 0
                    
                    # 计算其他希腊字母
                    T = calculate_days_to_expiry(expiry, self.trade_date)
                    greeks = calculate_greeks(self.underlying_price, strike, T, self.risk_free_rate, put_iv if put_iv > 0 else 0.2, 'P')
                    put_gamma = greeks['gamma']
                    put_theta = greeks['theta']
                    put_vega = greeks['vega']
                else:
                    put_code = ''
                    put_price = put_close = put_settlement = put_change1 = put_change2 = 0
                    put_iv = put_delta = put_gamma = put_theta = put_vega = 0
                    put_volume = put_oi = put_iv_change = put_oi_change = put_volume_change = put_exercise = 0
                    prev_put = {}
                
                all_strike_rows.append({
                    'expiry': expiry,
                    'actual_expiry_date': get_expiry_date(expiry),
                    'strike': float(strike),
                    'call_code': str(call_code),
                    'call_price': float(call_price),           # 今结算
                    'call_close': float(call_close),           # 今收盘
                    'call_settlement': float(call_settlement), # 昨结算
                    'call_change1': float(call_change1),      # 涨跌1
                    'call_change2': float(call_change2),      # 涨跌2
                    'call_iv': float(call_iv) * 100 if call_iv else 0,
                    'call_iv_change': float(call_iv_change),
                    'call_delta': float(call_delta),
                    'call_gamma': float(call_gamma),
                    'call_theta': float(call_theta),
                    'call_vega': float(call_vega),
                    'call_volume': int(call_volume),
                    'call_volume_change': int(call_volume_change),
                    'call_oi': int(call_oi),
                    'call_oi_change': float(call_oi_change),
                    'call_exercise': int(call_exercise),
                    'prev_call_iv': float(prev_call.get('iv', 0)) * 100 if prev_call else 0.0,
                    'put_code': str(put_code),
                    'put_price': float(put_price),            # 今结算
                    'put_close': float(put_close),              # 今收盘
                    'put_settlement': float(put_settlement),   # 昨结算
                    'put_change1': float(put_change1),        # 涨跌1
                    'put_change2': float(put_change2),        # 涨跌2
                    'put_iv': float(put_iv) * 100 if put_iv else 0,
                    'put_iv_change': float(put_iv_change),
                    'put_delta': float(put_delta),
                    'put_gamma': float(put_gamma),
                    'put_theta': float(put_theta),
                    'put_vega': float(put_vega),
                    'put_volume': int(put_volume),
                    'put_volume_change': int(put_volume_change),
                    'put_oi': int(put_oi),
                    'put_oi_change': float(put_oi_change),
                    'put_exercise': int(put_exercise),
                    'prev_put_iv': float(prev_put.get('iv', 0)) * 100 if prev_put else 0.0
                })
        
        # 按到期日（实际日期）、行权价排序
        all_strike_rows.sort(key=lambda x: (x['actual_expiry_date'], x['strike']))
        
        # 按到期日排序（按实际日期）
        expiry_list.sort(key=lambda x: x.actual_expiry_date)
        
        # 转换为StrikeRow列表
        strike_rows = [StrikeRow(**row) for row in all_strike_rows]
        
        return expiry_list, strike_rows
    
    def calculate_thresholds(self, iv: float) -> Dict[str, float]:
        """计算IV变化阈值
        
        Args:
            iv: 当前IV (%)
        """
        if iv < 20:
            return {'noise': IV_CHANGE_LOW, 'significant': IV_CHANGE_LOW * 2, 'extreme': IV_CHANGE_LOW * 3}
        elif iv < 30:
            return {'noise': IV_CHANGE_MID, 'significant': IV_CHANGE_MID * 2, 'extreme': IV_CHANGE_MID * 3}
        else:
            return {'noise': IV_CHANGE_HIGH, 'significant': IV_CHANGE_HIGH * 2, 'extreme': IV_CHANGE_HIGH * 3}

# ==================== 数据库存储 ====================

class OptionDataStore:
    """期权数据存储 (SQLite)"""
    
    def __init__(self, db_path: str = 'option_data.db'):
        self.db_path = db_path
        self._init_db()
    
    def _init_db(self):
        """初始化数据库"""
        conn = sqlite3.connect(self.db_path)
        c = conn.cursor()
        
        # 期权日线数据
        c.execute('''CREATE TABLE IF NOT EXISTS option_daily
                     (symbol TEXT, trade_date TEXT, last_price REAL, 
                      iv REAL, delta REAL, gamma REAL, theta REAL, vega REAL,
                      volume INTEGER, oi INTEGER, PRIMARY KEY (symbol, trade_date))''')
        
        # 期权Session快照数据（用于日内变化计算）
        c.execute('''CREATE TABLE IF NOT EXISTS option_session_snapshots
                     (symbol TEXT, trade_date TEXT, session_type TEXT,
                      last_price REAL, iv REAL, volume INTEGER, oi INTEGER,
                      created_at TEXT,
                      PRIMARY KEY (symbol, trade_date, session_type))''')
        
        # 历史波动锥数据
        c.execute('''CREATE TABLE IF NOT EXISTS vol_cone
                     (trade_date TEXT, tenor TEXT, hv REAL,
                      hv_min REAL, hv_25pct REAL, hv_median REAL, hv_75pct REAL, hv_max REAL,
                      PRIMARY KEY (trade_date, tenor))''')
        
        conn.commit()
        conn.close()
    
    def save_option_data(self, df: pd.DataFrame, trade_date: str):
        """保存期权数据"""
        if df is None or len(df) == 0 or not trade_date:
            return
        
        conn = sqlite3.connect(self.db_path)
        c = conn.cursor()
        
        for _, row in df.iterrows():
            code = row['合约代码']
            try:
                c.execute('''INSERT OR REPLACE INTO option_daily 
                             (symbol, trade_date, last_price, iv, delta, volume, oi)
                             VALUES (?, ?, ?, ?, ?, ?, ?)''',
                         (code, trade_date,
                          row.get('今结算') or row.get('今收盘'),
                          get_iv_from_data(row),
                          get_delta_from_data(row),
                          int(row.get('成交量(手)', 0) or 0),
                          int(row.get('持仓量', 0) or 0)))
            except:
                continue
        
        conn.commit()
        conn.close()
    
    def save_session_snapshot(self, df: pd.DataFrame, trade_date: str, session_type: str):
        """保存Session快照"""
        if df is None or len(df) == 0 or not trade_date:
            return
        conn = sqlite3.connect(self.db_path)
        c = conn.cursor()
        now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        for _, row in df.iterrows():
            code = row['合约代码']
            try:
                c.execute('''INSERT OR REPLACE INTO option_session_snapshots 
                             (symbol, trade_date, session_type, last_price, iv, volume, oi, created_at)
                             VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
                         (code, trade_date, session_type,
                          row.get('今结算') or row.get('今收盘'),
                          get_iv_from_data(row),
                          int(row.get('成交量(手)', 0) or 0),
                          int(row.get('持仓量', 0) or 0),
                          now_str))
            except:
                continue
        conn.commit()
        conn.close()
    
    def get_session_snapshot(self, trade_date: str, session_type: str) -> pd.DataFrame:
        """获取指定Session的快照数据"""
        conn = sqlite3.connect(self.db_path)
        c = conn.cursor()
        c.execute('''SELECT symbol, last_price, iv, volume, oi, session_type
                     FROM option_session_snapshots WHERE trade_date = ? AND session_type = ?''',
                   (trade_date, session_type))
        rows = []
        for row in c.fetchall():
            rows.append({
                '合约代码': row[0], 'last_price': row[1], 'iv': row[2],
                'volume': row[3], 'oi': row[4], 'session_type': row[5]
            })
        conn.close()
        return pd.DataFrame(rows) if rows else pd.DataFrame()
    
    def get_latest_session_close(self, trade_date: str = None) -> tuple:
        """获取最近一次Session收盘的数据"""
        if trade_date is None:
            trade_date = datetime.now().strftime('%Y%m%d')
        now = datetime.now()
        hour = now.hour + now.minute / 60
        if hour >= 15:
            session_type = 'afternoon'
        elif hour >= 11.5:
            session_type = 'morning'
        else:
            dt = datetime.strptime(trade_date, '%Y%m%d') - timedelta(days=1)
            trade_date = dt.strftime('%Y%m%d')
            session_type = 'afternoon'
        df = self.get_session_snapshot(trade_date, session_type)
        return session_type, df
    
    def get_prev_day_data(self, trade_date: str = None) -> pd.DataFrame:
        """获取昨日数据（自动跳过非交易日）"""
        if trade_date is None:
            trade_date = datetime.now().strftime('%Y%m%d')
        
        conn = sqlite3.connect(self.db_path)
        c = conn.cursor()
        
        # 检查数据库中有哪些交易日
        c.execute('SELECT DISTINCT trade_date FROM option_daily ORDER BY trade_date DESC')
        available_dates = [row[0] for row in c.fetchall()]
        
        # 找出比当前交易日早的最新交易日
        prev_date = None
        for d in available_dates:
            if d < trade_date:
                prev_date = d
                break
        
        if prev_date is None:
            conn.close()
            return pd.DataFrame()
        
        # 表结构: symbol, trade_date, last_price, iv, delta, gamma, theta, vega, volume, oi
        c.execute('''SELECT symbol, last_price, iv, volume, oi 
                     FROM option_daily WHERE trade_date = ?''', (prev_date,))
        
        rows = []
        for row in c.fetchall():
            rows.append({
                '合约代码': row[0],
                'last_price': row[1],
                'iv': row[2],
                'volume': row[3],
                'oi': row[4]
            })
        
        conn.close()
        return pd.DataFrame(rows) if rows else pd.DataFrame()

# ==================== 主API类 ====================

class OptionChainAPI:
    """期权链API"""
    
    def __init__(self):
        self.analyzer = OptionChainAnalyzer()
        self.store = OptionDataStore()
        self.trade_date = None
        self._lock = threading.Lock()
        self._cache = {}
        self._cache_ttl = 300  # 缓存5分钟
    
    def get_full_chain(self) -> Dict[str, Any]:
        """获取完整期权链数据"""
        import warnings
        warnings.filterwarnings('ignore')
        
        with self._lock:
            now = time.time()
            
            # 检查缓存
            if self._cache and self._last_update:
                if now - self._last_update < self._cache_ttl:
                    return self._cache
            
            # 获取交易日期 - 优先尝试今日，其次最近交易日
            today = datetime.now()
            weekday = today.weekday()
            
            trade_date = None
            
            # 首先尝试今日（收盘后一般就有当日数据）
            try:
                today_str = today.strftime('%Y%m%d')
                df_test = ak.option_hist_czce(symbol='PTA期权', trade_date=today_str)
                if df_test is not None and len(df_test) > 0:
                    trade_date = today_str
            except:
                pass
            
            # 如果今日无数据，尝试最近交易日
            if trade_date is None:
                if weekday == 6:  # Sunday
                    start_offset = 2
                elif weekday == 5:  # Saturday
                    start_offset = 1
                else:
                    start_offset = 1  # 今天之前最近交易日
                
                for i in range(start_offset, 15):
                    dt = today - timedelta(days=i)
                    date_str = dt.strftime('%Y%m%d')
                    try:
                        df_test = ak.option_hist_czce(symbol='PTA期权', trade_date=date_str)
                        if df_test is not None and len(df_test) > 0:
                            trade_date = date_str
                            break
                    except:
                        continue
            
            if trade_date is None:
                    return {'success': False, 'error': '无法获取交易日'}
            
            self.trade_date = trade_date
            
            # 获取当前期权数据（用于确定近月合约）
            df = AkshareOptionData.get_option_data(self.trade_date)
            if df is None or len(df) == 0:
                return {'success': False, 'error': '获取期权数据失败'}
            
            # 获取昨日数据用于IV日间对比
            prev_df = self.store.get_prev_day_data(self.trade_date)
            
            # 保存今日数据
            self.store.save_option_data(df, trade_date)
            
            # 第一步：先确定近月合约（不计算Greeks，只需要解析数据）
            # 按到期日分组获取expiry列表
            expiry_groups = df.groupby('expiry')
            temp_expiry_list = []
            for expiry, group in expiry_groups:
                temp_expiry_list.append({
                    'expiry': expiry,
                    'actual_expiry_date': get_expiry_date(expiry)
                })
            # 按实际到期日排序
            temp_expiry_list.sort(key=lambda x: x['actual_expiry_date'])
            near_expiry_code = temp_expiry_list[0]['expiry'] if temp_expiry_list else None
            
            # 第二步：获取真实TA606期货标的价格
            # 方法1：优先使用TqSdk获取CZCE.TA606的实时价格
            # 方法2：用Put-Call Parity从期权数据估算
            # 方法3：回退到akshare的TA0（主力合约）
            # 方法4：回退到成交量最大期权的行权价
            
            S = None
            
            # 方法1：用TqSdk获取TA606实时价格
            S = get_tq_ta606_price(timeout=5.0)
            
            # 方法2：用Put-Call Parity从期权数据估算
            if not S or S <= 0:
                if near_expiry_code:
                    S = estimate_underlying_from_options(df, near_expiry_code)
            
            # 方法3：回退到akshare的TA0（主力合约）
            if not S or S <= 0:
                try:
                    ta_df = ak.futures_zh_minute_sina(symbol='TA0', period='1m')
                    ta_df.columns = [c.strip() for c in ta_df.columns]
                    S = float(ta_df['close'].iloc[-1])
                except:
                    pass
            
            # 方法3：回退到成交量最大期权的行权价
            if S is None or S <= 0:
                try:
                    S = float(df.loc[df['成交量(手)'].idxmax(), 'strike'])
                except:
                    S = 6300  # 默认值
            
            # 第三步：设置analyzer的标的价格（在调用build_t_type_quote之前！）
            self.analyzer.underlying_price = S
            
            # 第四步：构建T型报价（此时underlying_price已正确设置）
            expiry_list, strike_rows = self.analyzer.build_t_type_quote(df, prev_df)
            
            if len(strike_rows) == 0:
                return {'success': False, 'error': '解析期权数据失败'}
            
            # 计算ATM行权价
            atm_strike = round(S / 50) * 50  # PTA最小跳动50
            
            # 计算统计数据
            stats = self._calculate_stats(expiry_list, strike_rows)
            
            # 构建返回
            # 只返回最近月合约数据
            near_strike_rows = [r for r in strike_rows if r.expiry == near_expiry_code]
            near_stats = self._calculate_stats([expiry_list[0]], near_strike_rows) if expiry_list else {}
            near_stats['full_market_pcr'] = self._calculate_full_market_pcr(expiry_list)
            
            result = {
                'success': True,
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'trade_date': self.trade_date,
                'underlying': near_expiry_code if near_expiry_code else 'CZCE.TA',  # 近月合约代码
                'underlying_price': S,
                'atm_strike': atm_strike,
                'near_expiry': near_expiry_code,  # 合约代码 TA606
                'near_expiry_date': expiry_list[0].actual_expiry_date if expiry_list else None,  # 实际到期日 20260513
                'near_expiry_display': f"{near_expiry_code}（到期日：{expiry_list[0].actual_expiry_date[:4]}/{expiry_list[0].actual_expiry_date[4:6]}/{expiry_list[0].actual_expiry_date[6:8]}）" if expiry_list else None,
                'expiry_list': [asdict(e) for e in expiry_list],  # 全部到期日列表
                'strike_rows': [asdict(r) for r in near_strike_rows],  # 只返回近月数据
                'stats': near_stats
            }
            
            self._cache = result
            self._last_update = now
            
            return result
    
    def _last_update(self):
        return getattr(self, '_last_update', None)
    
    def _calculate_stats(self, expiry_list: List[ExpiryData], 
                         strike_rows: List[StrikeRow]) -> Dict[str, Any]:
        """计算统计数据"""
        if not expiry_list or not strike_rows:
            return {}
        
        # 近月数据
        near_expiry = expiry_list[0]
        
        # ATM附近的IV
        atm_row = next((r for r in strike_rows if r.strike == self.analyzer.underlying_price), None)
        if atm_row:
            call_iv = atm_row.call_iv
            put_iv = atm_row.put_iv
        else:
            call_ivs = [r.call_iv for r in strike_rows if r.call_iv > 0]
            put_ivs = [r.put_iv for r in strike_rows if r.put_iv > 0]
            call_iv = sum(call_ivs) / len(call_ivs) if call_ivs else 0
            put_iv = sum(put_ivs) / len(put_ivs) if put_ivs else 0
        
        return {
            'near_expiry': near_expiry.expiry,
            'volume_pcr': near_expiry.volume_pcr,
            'position_pcr': near_expiry.position_pcr,
            'avg_call_iv': round(call_iv, 2),
            'avg_put_iv': round(put_iv, 2),
            'iv_diff': round(put_iv - call_iv, 2),
            'total_call_oi': near_expiry.total_oi_call,
            'total_put_oi': near_expiry.total_oi_put
        }
    
    def _calculate_full_market_pcr(self, expiry_list: List[ExpiryData]) -> Dict[str, float]:
        """计算全市场PCR"""
        total_call_vol = sum(e.total_volume_call for e in expiry_list)
        total_put_vol = sum(e.total_volume_put for e in expiry_list)
        total_call_oi = sum(e.total_oi_call for e in expiry_list)
        total_put_oi = sum(e.total_oi_put for e in expiry_list)
        return {
            'volume_pcr': round(total_put_vol / total_call_vol, 4) if total_call_vol > 0 else 0,
            'position_pcr': round(total_put_oi / total_call_oi, 4) if total_call_oi > 0 else 0
        }
    
    def get_volatility_cone(self) -> Dict[str, Any]:
        """获取波动率锥 - 使用akshare获取期货历史数据计算HV
        时间轴: 5、10、15、20、25、30、60、90、120日
        """
        import math
        
        try:
            # 步骤1: 检查数据库缓存
            conn = sqlite3.connect(self.store.db_path)
            c = conn.cursor()
            today = datetime.now().strftime('%Y%m%d')
            
            c.execute("SELECT tenor, hv, hv_min, hv_25pct, hv_median, hv_75pct, hv_max FROM vol_cone WHERE trade_date = ?", (today,))
            rows = c.fetchall()
            if rows and len(rows) >= 5:
                # 解析tenors为数字用于排序
                tenors_raw = [r[0] for r in rows]
                tenors_numeric = []
                for t in tenors_raw:
                    # 解析 "10日" -> 10
                    try:
                        tenors_numeric.append(int(t.replace('日', '')))
                    except:
                        tenors_numeric.append(0)
                
                # 按数字排序
                sorted_indices = sorted(range(len(tenors_numeric)), key=lambda i: tenors_numeric[i])
                tenors = [str(tenors_numeric[i]) + '日' for i in sorted_indices]
                
                # 按排序后的顺序重排数据
                hv_current = rows[sorted_indices[-1]][1] if sorted_indices else 20.0
                hv_min = [rows[i][2] for i in sorted_indices]
                hv_25pct = [rows[i][3] for i in sorted_indices]
                hv_median = [rows[i][4] for i in sorted_indices]
                hv_75pct = [rows[i][5] for i in sorted_indices]
                hv_max = [rows[i][6] for i in sorted_indices]
                
                # 计算HV百分位
                current_60_idx = tenors.index('60日') if '60日' in tenors else -1
                if current_60_idx >= 0 and hv_median[current_60_idx] > 0:
                    hv_25 = hv_25pct[current_60_idx]
                    hv_75 = hv_75pct[current_60_idx]
                    range_75 = hv_75 - hv_25
                    if range_75 > 0:
                        dist_from_25 = hv_current - hv_25
                        hv_percentile = 25 + (dist_from_25 / range_75) * 50
                        hv_percentile = max(0, min(100, hv_percentile))
                    else:
                        hv_percentile = 50.0
                else:
                    hv_percentile = 50.0
                
                # 从self获取ATM IV和到期天数
                try:
                    chain = self.get_full_chain()
                    atm_strike = chain.get('atm_strike', 0)
                    strike_rows = chain.get('strike_rows', [])
                    atm_row = next((r for r in strike_rows if r.get('strike') == atm_strike), None)
                    atm_call_iv = atm_row.get('call_iv', 0) if atm_row else 0
                    atm_put_iv = atm_row.get('put_iv', 0) if atm_row else 0
                    atm_iv = (atm_call_iv + atm_put_iv) / 2 if atm_call_iv or atm_put_iv else 0
                    near_expiry = chain.get('near_expiry', '')
                    if near_expiry:
                        days_to_exp = int(calculate_days_to_expiry(near_expiry, today) * 365)
                    else:
                        days_to_exp = 24
                except:
                    atm_iv = 0
                    days_to_exp = 24
                
                # 计算IV百分位（ATM IV在HV分布中的位置）
                iv_percentile = 50.0
                if atm_iv > 0 and current_60_idx >= 0:
                    hv_25 = hv_25pct[current_60_idx]
                    hv_75 = hv_75pct[current_60_idx]
                    range_75 = hv_75 - hv_25
                    if range_75 > 0:
                        dist_from_med = atm_iv - hv_median[current_60_idx]
                        dist_from_25 = atm_iv - hv_25
                        iv_percentile = 25 + (dist_from_25 / range_75) * 50
                        iv_percentile = max(0, min(100, iv_percentile))
                
                return {
                    'success': True,
                    'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'hv_current': hv_current,
                    'hv_percentile': round(hv_percentile, 1),
                    'atm_iv': atm_iv,
                    'iv_percentile': round(iv_percentile, 1),
                    'days_to_expiry': days_to_exp,
                    'tenors': tenors,
                    'hv_min': hv_min,
                    'hv_25pct': hv_25pct,
                    'hv_median': hv_median,
                    'hv_75pct': hv_75pct,
                    'hv_max': hv_max
                }
            conn.close()
            
            # 步骤2: 从akshare获取期货日线数据
            try:
                df = ak.futures_zh_daily_sina(symbol="TA0")
            except Exception as e:
                return {
                    'success': True,
                    'note': 'akshare数据获取失败: ' + str(e)[:50],
                    'hv_current': 0, 'tenors': [],
                    'hv_min': [], 'hv_25pct': [], 'hv_median': [], 'hv_75pct': [], 'hv_max': []
                }
            
            if df is None or len(df) < 120:
                return {
                    'success': True,
                    'note': '期货历史数据不足' + str(len(df) if df is not None else 0) + '天，需至少120个交易日',
                    'hv_current': 0, 'tenors': [],
                    'hv_min': [], 'hv_25pct': [], 'hv_median': [], 'hv_75pct': [], 'hv_max': []
                }
            
            # 提取收盘价（过滤NaN和无意义数据）
            df_clean = df.dropna(subset=['close'])
            df_clean = df_clean[df_clean['close'] > 0]
            closes = df_clean['close'].tolist()
            # 取最近的数据（确保数据是完整的）
            closes = closes[-500:]  # 只取最近500个交易日
            if len(closes) < 120:
                return {
                    'success': True,
                    'note': '有效数据不足' + str(len(closes)) + '天',
                    'hv_current': 0, 'tenors': [],
                    'hv_min': [], 'hv_25pct': [], 'hv_median': [], 'hv_75pct': [], 'hv_max': []
                }
            
            # 步骤3: 计算HV
            def calc_hv_series(prices, window):
                """计算滚动HV：基于每日对数收益率的滚动标准差年化"""
                if len(prices) < window + 1:
                    return []
                hvs = []
                # 先计算所有日收益率
                daily_returns = []
                for i in range(1, len(prices)):
                    if prices[i-1] > 0 and prices[i] > 0:
                        ret = math.log(prices[i] / prices[i-1])
                        daily_returns.append(ret)
                
                if len(daily_returns) < window:
                    return []
                
                # 滚动窗口计算HV
                for i in range(window, len(daily_returns)):
                    window_rets = daily_returns[i-window:i]
                    if len(window_rets) < window:
                        continue
                    mean = sum(window_rets) / len(window_rets)
                    variance = sum((r - mean) ** 2 for r in window_rets) / (len(window_rets) - 1)
                    hv = math.sqrt(variance * 252) * 100
                    hvs.append(hv)
                return hvs
            
            tenors = [5, 10, 15, 20, 25, 30, 60, 90, 120]
            tenors_display = [str(t) + '日' for t in tenors]
            
            hv_series_dict = {}
            for t in tenors:
                hv_series_dict[t] = calc_hv_series(closes, t)
            
            hv_min, hv_25pct, hv_median, hv_75pct, hv_max = [], [], [], [], []
            
            for t in tenors:
                hvs = hv_series_dict[t]
                if len(hvs) < 10:
                    hv_min.append(10.0)
                    hv_25pct.append(15.0)
                    hv_median.append(20.0)
                    hv_75pct.append(25.0)
                    hv_max.append(30.0)
                else:
                    sorted_hvs = sorted(hvs)
                    n = len(sorted_hvs)
                    hv_min.append(sorted_hvs[0])
                    hv_25pct.append(sorted_hvs[int(n * 0.25)])
                    hv_median.append(sorted_hvs[int(n * 0.50)])
                    hv_75pct.append(sorted_hvs[int(n * 0.75)])
                    hv_max.append(sorted_hvs[n - 1])
            
            current_hv = hv_series_dict[60][-1] if hv_series_dict[60] else 20.0
            
            current_60_hvs = hv_series_dict[60]
            if current_60_hvs:
                sorted_60 = sorted(current_60_hvs)
                rank = sum(1 for v in sorted_60 if v <= current_60_hvs[-1])
                hv_percentile = rank / len(sorted_60) * 100
            else:
                hv_percentile = 50.0
            
            # 步骤4: 保存到数据库缓存
            conn = sqlite3.connect(self.store.db_path)
            c = conn.cursor()
            for i, t in enumerate(tenors_display):
                c.execute("INSERT OR REPLACE INTO vol_cone (trade_date, tenor, hv, hv_min, hv_25pct, hv_median, hv_75pct, hv_max) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                    (today, t, current_hv if i == 5 else 0, hv_min[i], hv_25pct[i], hv_median[i], hv_75pct[i], hv_max[i]))
            conn.commit()
            conn.close()
            
            return {
                'success': True,
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'hv_current': round(current_hv, 2),
                'hv_percentile': round(hv_percentile, 1),
                'tenors': tenors_display,
                'hv_min': [round(v, 2) for v in hv_min],
                'hv_25pct': [round(v, 2) for v in hv_25pct],
                'hv_median': [round(v, 2) for v in hv_median],
                'hv_75pct': [round(v, 2) for v in hv_75pct],
                'hv_max': [round(v, 2) for v in hv_max]
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'hv_current': 0, 'tenors': [],
                'hv_min': [], 'hv_25pct': [], 'hv_median': [], 'hv_75pct': [], 'hv_max': []
            }


# ==================== 全局实例 ====================

_option_api = None
_option_api_lock = threading.Lock()

def get_option_api() -> OptionChainAPI:
    """获取API单例"""
    global _option_api
    with _option_api_lock:
        if _option_api is None:
            _option_api = OptionChainAPI()
        return _option_api


# ==================== Flask路由 ====================

def register_option_routes(app):
    """注册Flask路由"""
    from flask import Blueprint, jsonify
    
    option_bp = Blueprint('option', __name__, url_prefix='/api/option')
    
    @option_bp.route('/chain', methods=['GET'])
    def get_option_chain():
        """获取期权链T型报价"""
        api = get_option_api()
        result = api.get_full_chain()
        return jsonify(result)
    
    @option_bp.route('/stats', methods=['GET'])
    def get_option_stats():
        """获取期权统计数据"""
        api = get_option_api()
        result = api.get_full_chain()
        return jsonify(result.get('stats', {}))
    
    @option_bp.route('/vol_cone', methods=['GET'])
    def get_vol_cone():
        """获取波动率锥"""
        api = get_option_api()
        result = api.get_volatility_cone()
        return jsonify(result)
    
    @option_bp.route('/refresh', methods=['POST'])
    def refresh():
        """刷新数据"""
        api = get_option_api()
        api._cache = None
        api._last_update = None
        result = api.get_full_chain()
        return jsonify(result)
    
    app.register_blueprint(option_bp)
    
    return option_bp


# ==================== 测试 ====================

if __name__ == '__main__':
    import akshare as ak
    
    print("Option Chain API Test")
    print("=" * 50)
    
    api = OptionChainAPI()
    
    print("Fetching option chain...")
    result = api.get_full_chain()
    
    if result.get('success'):
        print(f"Timestamp: {result['timestamp']}")
        print(f"Trade Date: {result['trade_date']}")
        print(f"Underlying: {result['underlying']}")
        print(f"Underlying Price: {result['underlying_price']}")
        print(f"ATM Strike: {result['atm_strike']}")
        print(f"Expiry dates: {len(result['expiry_list'])}")
        print(f"Strike rows: {len(result['strike_rows'])}")
        
        stats = result.get('stats', {})
        print(f"\nStats:")
        print(f"  Near Expiry: {stats.get('near_expiry')}")
        print(f"  Volume PCR: {stats.get('volume_pcr')}")
        print(f"  Position PCR: {stats.get('position_pcr')}")
        print(f"  Avg Call IV: {stats.get('avg_call_iv')}%")
        print(f"  Avg Put IV: {stats.get('avg_put_iv')}%")
        print(f"  IV Diff: {stats.get('iv_diff')}%")
    else:
        print(f"Error: {result.get('error')}")
# ==================== Excel 导出 ====================

def export_atm_option_excel(output_dir: str = None, trade_date: str = None) -> dict:
    """导出平值期权上下十档的隐波、成交、持仓数据到 Excel
    
    Args:
        output_dir: 输出目录，默认 ~/.hermes/option_exports/
        trade_date: 交易日期，默认当天
        
    Returns:
        {'success': True, 'filepath': '...'}
    """
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    
    if output_dir is None:
        output_dir = os.path.expanduser("~/.hermes/option_exports")
    os.makedirs(output_dir, exist_ok=True)
    
    if trade_date is None:
        trade_date = datetime.now().strftime("%Y%m%d")
    
    # 获取数据
    api = get_option_api()
    result = api.get_full_chain()
    if not result.get('success'):
        return {'success': False, 'error': result.get('error', '获取数据失败')}
    
    # 计算 ATM 行权价
    underlying_price = result.get('underlying_price', 0)
    atm_strike = round(underlying_price / 50) * 50
    strike_rows = result.get('strike_rows', [])
    
    # ATM 上下各10档，镜像对称排列：ATM居中，高行权价在上，低行权价在下
    all_strikes = sorted(set(r['strike'] for r in strike_rows))  # 低→高
    atm_idx = None
    for i, s in enumerate(all_strikes):
        if abs(s - atm_strike) < 25:
            atm_idx = i
            break

    if atm_idx is None:
        return {'success': False, 'error': f'未找到 ATM 行权价 {atm_strike}'}

    start_idx = max(0, atm_idx - 10)
    end_idx = min(len(all_strikes), atm_idx + 11)
    selected_strikes = all_strikes[start_idx:end_idx]  # ATM在中间，高→低（因为逆序取片段）
    selected_strikes = list(reversed(selected_strikes))  # 翻转：让ATM在第11行，高Strike在上
    
    # 建立 strike -> row 映射
    row_map = {r['strike']: r for r in strike_rows}
    
    # 近月信息
    near_expiry = result.get('near_expiry', '')
    near_expiry_display = result.get('near_expiry_display', near_expiry)
    # sheet名不允许 / : * ? [ ] ，替换掉
    sheet_name = re.sub(r'[/:\\*?\[\]]', '-', near_expiry_display)
    # ---- 构建工作簿 ----
    wb = Workbook()

    # === 样式定义 ===
    hdr_font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    sub_fill = PatternFill("solid", fgColor="D6E4F0")
    alt_fill = PatternFill("solid", fgColor="EBF3FB")
    atm_fill = PatternFill("solid", fgColor="FFF2CC")
    atm_hdr_fill = PatternFill("solid", fgColor="FFD966")

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right = Alignment(horizontal='right', vertical='center')
    left = Alignment(horizontal='left', vertical='center')

    thin = Side(style='thin', color='BFBFBF')
    medium = Side(style='medium', color='4472C4')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    medium_border = Border(left=medium, right=medium, top=medium, bottom=medium)

    # 涨跌颜色
    up_font = Font(name='微软雅黑', color='C00000', size=10)
    dn_font = Font(name='微软雅黑', color='008000', size=10)
    neu_font = Font(name='微软雅黑', color='000000', size=10)
    bold_font = Font(name='微软雅黑', bold=True, size=10)
    norm_font = Font(name='微软雅黑', size=10)

    def chg_font(val):
        if val is None: return neu_font
        return up_font if val > 0 else dn_font if val < 0 else neu_font

    def chg_sign(val):
        if val is None: return '0.0%'
        sign = '+' if val > 0 else ''
        return f'{sign}{val:.1f}%'

    def fill_font_and_align(ws, row, col, value, font, align=None):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = font
        cell.alignment = align or center
        cell.border = thin_border
        return cell

    # === Sheet: 认购+认沽合并 ===
    ws = wb.active
    ws.title = sheet_name

    # 主标题
    ws.merge_cells(f'A1:M1')
    title_cell = ws['A1']
    title_cell.value = f'PTA期权链 (ATM±10档)  {near_expiry_display}  {trade_date}'
    title_cell.font = Font(name='微软雅黑', bold=True, size=13, color='FFFFFF')
    title_cell.fill = PatternFill("solid", fgColor="1F4E79")
    title_cell.alignment = center
    ws.row_dimensions[1].height = 28

    # 副标题
    ws.merge_cells(f'A2:M2')
    sub = ws['A2']
    sub.value = f'标的: PTA  当前价: {underlying_price:.0f}  ATM行权价: {atm_strike:.0f}'
    sub.font = Font(name='微软雅黑', size=10, color='595959', italic=True)
    sub.fill = PatternFill("solid", fgColor="D9E1F2")
    sub.alignment = center
    ws.row_dimensions[2].height = 18

    # 列标题：镜像对称排列（从行权价向两侧展开，越近越靠内）
    # Call区（左侧6列）：持仓变化 | 持仓量 | 成交量变化 | 成交量 | IV变化 | IV
    # Put区（右侧6列）：IV | IV变化 | 成交量 | 成交量变化 | 持仓量 | 持仓变化
    col_headers = [
        'Call持仓变化', 'Call持仓量', 'Call成交量变化', 'Call成交量', 'Call-IV变化(%)', 'Call-IV(%)',
        '行权价',
        'Put-IV(%)', 'Put-IV变化(%)', 'Put成交量', 'Put成交量变化', 'Put持仓量', 'Put持仓变化'
    ]

    # Call区 header（绿色）
    hdr_call_fill = PatternFill("solid", fgColor="375623")
    for ci, h in enumerate(col_headers[:6], 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = hdr_font
        c.fill = hdr_call_fill
        c.alignment = center
        c.border = thin_border

    # 行权价列（ATM高亮）
    c_strike = ws.cell(row=3, column=7, value='行权价')
    c_strike.font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
    c_strike.fill = PatternFill("solid", fgColor="2F5496")
    c_strike.alignment = center
    c_strike.border = thin_border

    # Put区 header（蓝色）
    hdr_put_fill = PatternFill("solid", fgColor="1F3864")
    for ci, h in enumerate(col_headers[7:], 8):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = hdr_font
        c.fill = hdr_put_fill
        c.alignment = center
        c.border = thin_border

    ws.row_dimensions[3].height = 22

    norm_font = Font(name='微软雅黑', size=10)
    bold_font = Font(name='微软雅黑', bold=True, size=10)
    call_fill_even = PatternFill("solid", fgColor="E2EFDA")
    call_fill_odd = PatternFill("solid", fgColor="EBF3E8")
    put_fill_even = PatternFill("solid", fgColor="D6E4F0")
    put_fill_odd = PatternFill("solid", fgColor="EBF3FB")
    atm_fill = PatternFill("solid", fgColor="FFF2CC")
    atm_strike_fill = PatternFill("solid", fgColor="FFD966")
    white_fill = PatternFill("solid", fgColor="FFFFFF")

    up_font = Font(name='微软雅黑', color='C00000', size=10)
    dn_font = Font(name='微软雅黑', color='008000', size=10)
    neu_font = Font(name='微软雅黑', color='000000', size=10)

    def chg_color(val):
        if val is None: return neu_font
        return up_font if val > 0 else dn_font if val < 0 else neu_font

    for ri, strike in enumerate(selected_strikes):
        r = row_map.get(strike)
        if not r:
            continue
        excel_row = ri + 4
        is_atm = abs(strike - atm_strike) < 25
        bg_c = atm_fill if is_atm else (call_fill_even if ri % 2 == 0 else call_fill_odd)
        bg_p = atm_fill if is_atm else (put_fill_even if ri % 2 == 0 else put_fill_odd)
        bg_s = atm_strike_fill if is_atm else PatternFill("solid", fgColor="D9E1F2")

        # Call 6列（从外向内：持仓变化→持仓量→成交量变化→成交量→IV变化→IV）
        call_vals = [
            r.get('call_oi_change'),
            r.get('call_oi'),
            r.get('call_volume_change'),
            r.get('call_volume'),
            r.get('call_iv_change'),
            r.get('call_iv'),
        ]
        # Put 6列（从左向右：IV→IV变化→成交量→成交量变化→持仓量→持仓变化）
        put_vals = [
            r.get('put_iv'),
            r.get('put_iv_change'),
            r.get('put_volume'),
            r.get('put_volume_change'),
            r.get('put_oi'),
            r.get('put_oi_change'),
        ]

        # Call 6列
        for ci, val in enumerate(call_vals, 1):
            cell = ws.cell(row=excel_row, column=ci, value=val)
            cell.fill = bg_c
            cell.alignment = center
            cell.border = thin_border
            # ci=1持仓变化/ci=3成交量变化/ci=5IV变化 用颜色
            if ci in (1, 3, 5):
                cell.font = chg_color(val)
                if val is not None:
                    cell.number_format = '+0.0%;-0.0%'
            elif ci in (2, 4):
                cell.font = norm_font
                if val is not None:
                    cell.number_format = '#,##0'
            elif ci == 6:
                cell.font = norm_font
                if val is not None:
                    cell.number_format = '0.0'

        # 行权价
        cell_s = ws.cell(row=excel_row, column=7, value=strike)
        cell_s.fill = bg_s
        cell_s.alignment = center
        cell_s.border = thin_border
        cell_s.font = bold_font

        # Put 6列
        for ci, val in enumerate(put_vals, 8):
            cell = ws.cell(row=excel_row, column=ci, value=val)
            cell.fill = bg_p
            cell.alignment = center
            cell.border = thin_border
            # Put: ci=8IV/ci=9IV变化/ci=10成交量/ci=11成交量变化/ci=12持仓量/ci=13持仓变化
            if ci == 8:
                cell.font = norm_font
                if val is not None:
                    cell.number_format = '0.0'
            elif ci == 9:
                cell.font = chg_color(val)
                if val is not None:
                    cell.number_format = '+0.0%;-0.0%'
            elif ci == 10:
                cell.font = norm_font
                if val is not None:
                    cell.number_format = '#,##0'
            elif ci == 11:
                cell.font = chg_color(val)
                if val is not None:
                    cell.number_format = '+0.0%;-0.0%'
            elif ci == 12:
                cell.font = norm_font
                if val is not None:
                    cell.number_format = '#,##0'
            elif ci == 13:
                cell.font = chg_color(val)
                if val is not None:
                    cell.number_format = '+0.0%;-0.0%'

        ws.row_dimensions[excel_row].height = 18

    col_widths = [14, 12, 14, 12, 14, 12, 12, 12, 14, 12, 14, 12, 14]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = 'A4'

    filename = f'PTA_option_chain_{near_expiry}_{trade_date}.xlsx'
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    
    return {'success': True, 'filepath': filepath, 'filename': filename}
